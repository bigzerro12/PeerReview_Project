import os.path
import pandas as pd
from openpyxl import workbook, load_workbook
import openpyxl


filepath =  "D:\Py\Project\PeerReview1\Scores.xlsx"
#1. showlist(), 2. total(), 3. max(), 4. min(), 5. create(), 6. addstudent(), 7.removestudent(), 8. convertstudent().
# checkIfMatchID Remove, checkfile emtpy, main
def Create():
  wb = openpyxl.Workbook()
  wb.save(filepath)
  ws = wb.active
  ws.title = "Data"
  ws['A1'].value = "ID"
  ws['B1'].value = "Name"
  ws['C1'].value = "Math"
  ws['D1'].value = "Physics"
  ws['E1'].value = "English"
  ws['F1'].value = "Average"
  wb.save(filepath)
#PRM_37G_001


# Create file function
def createEXFile():
  if (os.path.isfile(filepath) == 0):
    Create()
  else:
    print("WRN_001: This file already exists. Do you want to replace it?")
    print("Yes -> 1             No -> 2")
    a = 3
    while (a != 1 and a != 2):
      a = input("Enter your choice: ")
      a = int(a)
      if (a == 1):
        print("Are you sure?  Yes -> 1    No -> 2")
        b = input("Enter your choice: ")
        b = int(b)
        if (b == 1):
          os.remove(filepath)
          Create()
          print("The Students Score File has been recreated!")
#PRM_37G_002



def checkIfMatchID(ID):
  wb = load_workbook(filepath)
  ws = wb.active
  maxRow = ws.max_row

  for i in range (2, maxRow + 1):
    if (ws.cell(row = i, column = 1).value == ID):
      return 1
#PRM_37G_003

def is_float(string):
    try:
        float(string)
        return True
    except ValueError:
        return False
#PRM_37G_004



      
# Add new student function
def addStudent():

  wb = load_workbook(filepath)
  ws = wb.active

  while (True):
    studentName = input("Enter your student name: ")
    if (all(x.isalpha() or x.isspace() for x in studentName)):
      break
    else: 
      print("ERR_001: The student name invalid due to constains special characters, please use the different name.")

  while (True):
    studentID = input("Enter your student ID: ")
    if (studentID.isnumeric()):
      studentID = int(studentID)
      if (studentID <= 9999 and studentID >= 0):
        break
      else:
        print("ERR_002: ID is not valid, please try again with the integer ID in range [0 - 9999].")
    else:
      print("ERR_003: ID is not valid, please try again with the integer ID [1...].")


  if (checkIfMatchID(studentID) != 1):

    while (True):

      mathScore = input("Enter your student Math Score: ")
      physicsScore = input("Enter your student Physics Score: ")
      englishScore = input("Enter your student English Score: ")

      if (is_float(mathScore) and is_float(physicsScore) and is_float(englishScore)):
        floatMathScore = float(mathScore)
        floatPhysicsScore = float(physicsScore)
        floatEnglishScore = float(englishScore)

        if ((floatMathScore <= 10) and (floatMathScore >= 0) and (floatPhysicsScore <= 10) and (floatPhysicsScore >= 0) and (floatEnglishScore <= 10) and (floatEnglishScore >= 0)):
          break
        else:
          print("ERR_004: Score number is not valid, please try again with the float number in range [0 - 10].")
      else:
        print("ERR_005: Score number is not valid, please try again with the float number.")

    floatAverage = (floatMathScore + floatPhysicsScore + floatEnglishScore) / 3
    Average = repr(floatAverage)
    
    ws.append([studentID, studentName, mathScore, physicsScore, englishScore, Average])
    wb.save(filepath)
    print("Added Successfully!")
    showList()
  else:
    print("ERR_006: ID already exists!")
#PRM_37G_005



    
# Remove function
def Remove():
  wb = load_workbook(filepath)
  ws = wb.active

  maxRow = ws.max_row
  temp = 1

  ID = input("Enter student ID you want to delete: ")

  if (ID.isnumeric()):
    ID = int(ID)

    for i in range (2, maxRow + 1):
      print(ws.cell(row = i, column = 1).value)
      print(i)
      if (ws.cell(row = i, column = 1).value == ID):
        temp = 0
        print("WRN_002: Are you sure?  Yes -> 1    No -> 2")
        a = input("Enter your choice: ")
        a = int(a)
        if (a == 1):
          ws.delete_rows(idx = i)
          wb.save(filepath)
          print("Removed Successfully!")
          showList()
        break
        

    if (temp == 1):
      print("ERR_007: ID does not exist!")
  
  else:
    print("ERR_008: ID is not valid, please try again with the integer ID [1...].")
#PRM_37G_006

def checkIfFileEmpty():
  wb = load_workbook(filepath)
  ws = wb.active
  maxRow = ws.max_row

  if (maxRow == 1):

    return 1
#PRM_37G_007



# Max function
def Max():
  if (checkIfFileEmpty() != 1):
    df = pd.read_excel(filepath)
    maxAverageVal = df["Average"].max()
    maxAverageIndex = df["Average"].idxmax()

    maxAverageVal = repr(maxAverageVal)
    print("The highest average score: " + maxAverageVal)
    ID = df.values[maxAverageIndex][0]
    ID = repr(ID)
    print("ID of the Student has the highest average score: " + ID)
  else:
    print("The Data File is Empty!")
#PRM_37G_008




# Min function
def Min():
  if (checkIfFileEmpty() != 1):
    df = pd.read_excel(filepath)
    minAverageVal = df["Average"].min()
    minAverageIndex = df["Average"].idxmin()

    minAverageVal = repr(minAverageVal)
    print("The lowest average score: " + minAverageVal)
    ID = df.values[minAverageIndex][0]
    ID = repr(ID)
    print("ID of the Student has the lowest average score: " + ID)
  else:
    print("The Data File is Empty!")
#PRM_37G_009



# Total function 
def Total():
  wb = load_workbook(filepath)
  ws = wb.active
  maxRow = ws.max_row

  totalStudents = maxRow - 1
  totalStudents = repr(totalStudents) 

  print("The total number of students is: " + totalStudents)

def showList():
  df = pd.read_excel(filepath)
  print(df)
#PRM_37G_010

def showInstructionList():
  print("----------------------Options-----------------------")
  print("Create New File -> 1 ")
  print("Add new Student -> 2 ")
  print("Remove Student -> 3 ")
  print("Get ID of the Student has the highest average score -> 4 ")
  print("Get ID of the Student has the lowest average score -> 5 ")
  print("Show Total number of students -> 6")
  print("Show Students List -> 7")
  print("Turn OFF! -> 0")
#PRM_37G_011

def Exit():
  print("------------------------END--------------------------")


# ---------- Main function ------------
def main():

  if (os.path.isfile(filepath) == 0):
    Create()

  i = 9
  while(i != 0):
    
    showInstructionList()

    i = input("Enter your option: ")
    print("====================================================")

    if (i.isnumeric()):
      i = int(i)

      try:

        if (i == 1):
          createEXFile()
        elif (i == 2):
          addStudent()
        elif (i == 3):
          Remove()
        elif (i == 4):
          Max()
        elif (i == 5):
          Min()
        elif (i == 6):
          Total()
        elif (i == 7):
          showList()
        elif (i == 0):
          Exit()
          break
        else:
          print("ERR_009: Selection is not valid, please try again with the integer number in range [0 - 7].")

      except:
        print("Something went wrong. The program did not run properly!")

    else:
      print("ERR_010: Selection is not valid, please try again with the integer number.")
#PRM_37G_012
    


if __name__ == "__main__":
  main()
